from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from database import replace_merchant_mcc


EXPECTED_HEADERS = {
    "merchant full name": "merchant_full_name",
    "merchant name": "merchant_name",
    "address": "address",
    "amex": "amex",
    "payment method": "payment_method",
    "mcc": "mcc",
    "note": "note",
}


def import_file(path: str) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [str(value or "").strip().lower() for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index_to_field = {idx: EXPECTED_HEADERS.get(header) for idx, header in enumerate(headers)}

    rows = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {}
        for idx, value in enumerate(values):
            field = index_to_field.get(idx)
            if field:
                row[field] = str(value).strip() if value is not None else None
        rows.append(row)
    return replace_merchant_mcc(rows, source=Path(path).name)


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: python import_mcc_excel.py <path-to-xlsx>")
        return 1
    count = import_file(sys.argv[1])
    print(f"Imported {count} merchant MCC rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
